#!/usr/bin/env python
"""Analyze GO_POSITION and TURN durations within a ZPATH log interval."""

from __future__ import print_function

import argparse
import csv
import re
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path
from statistics import mean, median


DEFAULT_START = "15:57:29.918"
DEFAULT_END = "16:38:05.550"
DEFAULT_LOG_SOURCE = Path("log")
DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_EXCEL = DEFAULT_OUTPUT_DIR / "zpath_timing.xlsx"
DEFAULT_CHART_SAVE = DEFAULT_OUTPUT_DIR / "chart.png"
DEFAULT_ZPATH_CHART_SAVE = DEFAULT_OUTPUT_DIR / "zpath_chart.png"
DEFAULT_TASK_SAVE = DEFAULT_OUTPUT_DIR / "task_summary.png"
DEFAULT_CUT_PROGRESS_SAVE = DEFAULT_OUTPUT_DIR / "cut_progress.png"

LOG_PATTERN = re.compile(
    r"^\[[A-Z]\s+(?P<date>\d{2}-\d{2}-\d{2})\s+"
    r"(?P<time>\d{2}:\d{2}:\d{2}\.\d{3})\b.*?\]"
    r"HSM_(?P<level>\d+)_CUTTING_STATE_MACHINE:\s+"
    r"(?P<action>Inner|Sibling|Pop)\s+:\s+(?P<state>[A-Z0-9_]+)\s*$"
)

RAW_TIMESTAMP_PATTERN = re.compile(
    r"^\[[A-Z]\s+(?P<date>\d{2}-\d{2}-\d{2})\s+"
    r"(?P<time>\d{2}:\d{2}:\d{2}\.\d{3})\b"
)

GENERIC_HSM_PATTERN = re.compile(
    r"^\[[A-Z]\s+(?P<date>\d{2}-\d{2}-\d{2})\s+"
    r"(?P<time>\d{2}:\d{2}:\d{2}\.\d{3})\b.*?\]"
    r"HSM_(?P<level>\d+)_(?P<machine>[A-Z]+)_STATE_MACHINE:\s+"
    r"(?P<action>Init|Inner|Sibling|Pop)\s+:\s+(?P<state>[A-Z0-9_]+)\s*$"
)

POINT_TYPE_PATTERN = re.compile(
    r"point type:\s*(?P<type>[012])\(0:\s*astar\s+1:\s*z long\s+2:\s*z short\)"
)

CUT_INFO_PATTERN = re.compile(
    r"cut info time\s*\[\s*"
    r"(?P<cut_time>[-+]?\d+(?:\.\d+)?)\s+"
    r"(?P<total_time>[-+]?\d+(?:\.\d+)?)\s*\]\s*"
    r"area\s*\[\s*"
    r"(?P<cut_area>[-+]?\d+(?:\.\d+)?)\s+"
    r"(?P<total_area>[-+]?\d+(?:\.\d+)?)\s*\]\s*"
    r"total percent\[(?P<percent>[-+]?\d+(?:\.\d+)?)\]"
)

EDGE_TYPE_NAMES = {
    0: "ASTAR",
    1: "LONG_EDGE",
    2: "SHORT_EDGE",
    None: "INCOMPLETE",
}

# Keep the task-level metric vocabulary in one place.  The console summary,
# Excel export, and task overview chart deliberately use these same keys and
# definitions so that a "total" never silently means active COVERAGE time.
TASK_METRIC_DEFINITIONS = (
    ("cycle_total", "总时间（cycle_total，含充电）",
     "Total cycle wall-clock duration using the task start and task end; includes observed charging dwell.",
     "#303030"),
    ("coverage", "覆盖运行累计时长（仅 COVERAGE 运行态）",
     "Sum of COVERAGE intervals; this is active runtime, not total cycle time.",
     "#4C4C4C"),
    ("edge", "沿边切割时间（EDGE_COVERAGE）",
     "Sum of EDGE_COVERAGE intervals inside COVERAGE.", "#E6AB02"),
    ("recharge", "回充动作累计（不含充电）",
     "Sum of HSM_0 DOCK_STATE_MACHINE DOCK intervals (Init to Pop); charging dwell is separate.",
     "#8172B2"),
    ("goto_breakpoint", "断点/回桩路径时间（DOCK.GOTO_ROAD）",
     "Sum of HSM_2 DOCK_STATE_MACHINE GOTO_ROAD intervals; separate from ZPATH.",
     "#55A868"),
    ("charging", "充电时间",
     "Explicitly observed charging dwell from DOCK_CHARGE pop until CUT resumes.",
     "#E1E1E1"),
    ("zpath", "弓字形切割总时间（长边、短边、ASTAR）",
     "GO path duration of completed, explicitly classified LONG_EDGE, SHORT_EDGE, and ASTAR traversals.",
     "#176B87"),
)

TASK_METRIC_CHART_LABELS = {
    "cycle_total": "Cycle total (incl. charging)",
    "coverage": "Coverage active runtime",
    "edge": "Edge cutting",
    "recharge": "Recharge movement excl. charging",
    "goto_breakpoint": "DOCK.GOTO_ROAD",
    "charging": "Charging",
    "zpath": "ZPATH classified cutting",
}

TASK_METRIC_CHART_KEYS = (
    "cycle_total", "coverage", "edge", "recharge", "goto_breakpoint",
    "charging", "zpath"
)
NON_CHARGING_TASK_METRIC_CHART_KEYS = tuple(
    key for key in TASK_METRIC_CHART_KEYS if key != "charging"
)
ZPATH_TASK_METRIC_CHART_KEYS = tuple(
    key for key in TASK_METRIC_CHART_KEYS
    if key not in ("charging", "goto_breakpoint")
)


def chart_metric_label(key):
    if key == "cycle_total":
        return "Cycle total excl. charging"
    return TASK_METRIC_CHART_LABELS[key]


def chart_metric_seconds(metrics, key):
    if key == "cycle_total":
        return metrics["cycle_total"] - metrics["charging"]
    return metrics[key]

MERGE_COVERAGE_GAP_SECONDS = 60.0


def parse_args():
    parser = argparse.ArgumentParser(
        description="统计 ZPATH 中 GO_POSITION 及其 TURN 的时长并绘图。"
    )
    parser.add_argument(
        "--log", default=str(DEFAULT_LOG_SOURCE),
        help="日志文件或日志目录路径（默认: log）",
    )
    parser.add_argument("--start", default=DEFAULT_START,
                        help="开始时间，HH:MM:SS.mmm 或 YY-MM-DD HH:MM:SS.mmm")
    parser.add_argument("--end", default=DEFAULT_END,
                        help="结束时间，HH:MM:SS.mmm 或 YY-MM-DD HH:MM:SS.mmm")
    parser.add_argument(
        "--save", metavar="PNG", default=str(DEFAULT_ZPATH_CHART_SAVE),
        help="将 ZPATH 图表保存为 PNG（默认: outputs/zpath_chart.png）",
    )
    parser.add_argument(
        "--task-save", metavar="PNG",
        default=str(DEFAULT_TASK_SAVE),
        help="生成日志内可观测覆盖任务的六项总览图（默认: outputs/task_summary.png）",
    )
    parser.add_argument(
        "--cut-save", metavar="PNG",
        default=str(DEFAULT_CUT_PROGRESS_SAVE),
        help="生成切割面积和切割百分比趋势图（默认: outputs/cut_progress.png）",
    )
    parser.add_argument(
        "--task-summary", action="store_true",
        help="打印日志内可观测覆盖任务的六项汇总",
    )
    parser.add_argument("--csv", metavar="CSV", help="导出逐次统计明细")
    parser.add_argument(
        "--excel", metavar="XLSX", default=str(DEFAULT_EXCEL),
        help="导出 Excel 汇总、明细和可观测任务指标（默认: outputs/zpath_timing.xlsx）",
    )
    parser.add_argument("--no-show", action="store_true", help="不弹出图表窗口")
    return parser.parse_args()


def timestamp_from_match(match):
    return datetime.strptime(
        match.group("date") + " " + match.group("time"),
        "%y-%m-%d %H:%M:%S.%f",
    )


def log_files(source):
    if source.is_dir():
        files = sorted(path for path in source.iterdir() if path.is_file())
        if not files:
            raise ValueError("日志目录为空: {}".format(source))
        return files
    if source.is_file():
        return [source]
    raise FileNotFoundError("日志文件或目录不存在: {}".format(source))


def load_log_records(source):
    """Load one file or a directory of rotated logs as one time-ordered stream."""
    records = []
    for file_index, path in enumerate(log_files(source)):
        with path.open("r", encoding="utf-8", errors="replace") as log_file:
            for source_line, text in enumerate(log_file, 1):
                timestamp_match = RAW_TIMESTAMP_PATTERN.match(text)
                if not timestamp_match:
                    continue
                records.append({
                    "time": timestamp_from_match(timestamp_match),
                    "text": text.rstrip(),
                    "source": str(path),
                    "source_index": file_index,
                    "source_line": source_line,
                })
    if not records:
        raise ValueError("日志中没有找到有效时间戳: {}".format(source))

    records.sort(key=lambda item: (
        item["time"], item["source_index"], item["source_line"]
    ))
    for line_number, record in enumerate(records, 1):
        record["line"] = line_number
    return records


def parse_log(records):
    events = []
    for record in records:
        match = LOG_PATTERN.match(record["text"])
        if not match:
            continue
        events.append({
            "time": record["time"],
            "level": int(match.group("level")),
            "action": match.group("action"),
            "state": match.group("state"),
            "line": record["line"],
        })
    if not events:
        raise ValueError("日志中没有找到 CUTTING_STATE_MACHINE 状态事件")
    return events


def parse_point_types(records):
    """Read explicit Z-path destination types from the raw log."""
    point_types = []
    for record in records:
        match = POINT_TYPE_PATTERN.search(record["text"])
        if match:
            point_types.append({
                "line": record["line"],
                "point_type": int(match.group("type")),
            })
    return point_types


def parse_cut_progress(records):
    """Read cut area and total percent samples from cut info lines."""
    progress = []
    for record in records:
        match = CUT_INFO_PATTERN.search(record["text"])
        if not match:
            continue
        cut_area = float(match.group("cut_area"))
        total_area = float(match.group("total_area"))
        area_percent = cut_area / total_area * 100.0 if total_area else 0.0
        progress.append({
            "time": record["time"],
            "line": record["line"],
            "cut_time": float(match.group("cut_time")),
            "total_time": float(match.group("total_time")),
            "cut_area": cut_area,
            "total_area": total_area,
            "area_percent": area_percent,
            "logged_percent": float(match.group("percent")),
        })
    return progress


def latest_true_zero_cut_progress(progress):
    """Return the first sample in the latest true zero-percent run."""
    latest_run_start = None
    in_zero_run = False
    for item in progress:
        is_true_zero = (
            item["total_area"] > 0
            and abs(item["cut_area"]) < 1e-9
            and abs(item["area_percent"]) < 1e-9
        )
        if is_true_zero and not in_zero_run:
            latest_run_start = item
        in_zero_run = is_true_zero
    return latest_run_start


def parse_raw_context(records):
    """Read the first timestamp plus CUTTING/DOCK HSM events."""
    raw_start = records[0]["time"] if records else None
    hsm_events = []
    for record in records:
        match = GENERIC_HSM_PATTERN.match(record["text"])
        if not match:
            continue
        hsm_events.append({
            "time": record["time"],
            "level": int(match.group("level")),
            "machine": match.group("machine"),
            "action": match.group("action"),
            "state": match.group("state"),
            "line": record["line"],
        })
    if raw_start is None:
        raise ValueError("日志中没有找到有效时间戳")
    return raw_start, hsm_events


def resolve_boundary(value, reference_date):
    for fmt in ("%y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S.%f"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    try:
        clock = datetime.strptime(value, "%H:%M:%S.%f").time()
    except ValueError:
        raise ValueError("时间格式错误: {}".format(value))
    return datetime.combine(reference_date, clock)


def has_explicit_date(value):
    return bool(re.match(r"^\d{2,4}-\d{2}-\d{2}\s+", value))


def candidate_windows(start_value, end_value, events):
    if has_explicit_date(start_value) or has_explicit_date(end_value):
        reference_date = events[0]["time"].date()
        start = resolve_boundary(start_value, reference_date)
        end = resolve_boundary(end_value, reference_date)
        if end <= start:
            end += timedelta(days=1)
        return [(start, end)]

    windows = []
    for reference_date in sorted({event["time"].date() for event in events}):
        start = resolve_boundary(start_value, reference_date)
        end = resolve_boundary(end_value, reference_date)
        if end <= start:
            end += timedelta(days=1)
        windows.append((start, end))
    return windows


def build_intervals(events, level, state, window_start, window_end,
                    infer_initial=False):
    """Build state intervals; a Sibling event replaces the state at its level."""
    intervals = []
    active = None
    saw_level_event = False

    for event in events:
        if event["level"] != level:
            continue
        action = event["action"]
        event_state = event["state"]

        # A log may begin while a state is already active.  If the first event
        # at this level is that state's Pop, retain the observable portion and
        # mark its start line as unknown instead of silently dropping it.
        if (infer_initial and not saw_level_event and action == "Pop"
                and event_state == state):
            synthetic_start = dict(event)
            synthetic_start["time"] = window_start
            synthetic_start["line"] = None
            active = synthetic_start
        saw_level_event = True

        if action == "Sibling":
            if active is not None:
                intervals.append((active, event))
                active = None
            if event_state == state:
                active = event
        elif action in ("Init", "Inner") and event_state == state:
            if active is not None:
                intervals.append((active, event))
            active = event
        elif action == "Pop" and event_state == state and active is not None:
            intervals.append((active, event))
            active = None

    if active is not None:
        synthetic_end = dict(active)
        synthetic_end["time"] = window_end
        synthetic_end["line"] = None
        intervals.append((active, synthetic_end))

    clipped = []
    for enter, leave in intervals:
        start = max(enter["time"], window_start)
        end = min(leave["time"], window_end)
        if end > start:
            clipped.append({
                "start": start,
                "end": end,
                "duration": (end - start).total_seconds(),
                "start_line": enter["line"],
                "end_line": leave["line"],
            })
    return clipped


def intersection_seconds(first, second):
    start = max(first["start"], second["start"])
    end = min(first["end"], second["end"])
    return max(0.0, (end - start).total_seconds())


def inside_any(interval, containers):
    return any(intersection_seconds(interval, container) > 0 for container in containers)


def clip_to_containers(intervals, containers):
    """Keep only the portions of intervals that are inside parent containers."""
    clipped = []
    for interval in intervals:
        for container in containers:
            start = max(interval["start"], container["start"])
            end = min(interval["end"], container["end"])
            if end <= start:
                continue
            clipped.append({
                "start": start,
                "end": end,
                "duration": (end - start).total_seconds(),
                "start_line": (
                    interval["start_line"]
                    if start == interval["start"] else container["start_line"]
                ),
                "end_line": (
                    interval["end_line"]
                    if end == interval["end"] else container["end_line"]
                ),
            })
    return clipped


def state_at_event(events, level, timestamp, line):
    """Return the active state at one HSM level before a child event occurs."""
    active_state = None
    for event in events:
        if event["time"] > timestamp or (
                line is not None and event["time"] == timestamp
                and event["line"] > line):
            break
        if event["level"] != level:
            continue
        if event["action"] in ("Init", "Inner", "Sibling"):
            active_state = event["state"]
        elif event["action"] == "Pop" and event["state"] == active_state:
            active_state = None
    return active_state or "UNKNOWN"


def summarize_goto_road_by_parent(hsm_events, window_start, window_end):
    """Aggregate every HSM_2 GOTO_ROAD interval by machine and HSM_1 parent."""
    grouped = {}
    machines = sorted({event["machine"] for event in hsm_events})
    for machine in machines:
        machine_events = [
            event for event in hsm_events if event["machine"] == machine
        ]
        intervals = build_intervals(
            machine_events, 2, "GOTO_ROAD", window_start, window_end,
            infer_initial=True,
        )
        for interval in intervals:
            parent_state = state_at_event(
                machine_events, 1, interval["start"], interval["start_line"]
            )
            key = (machine, parent_state)
            grouped.setdefault(key, []).append(interval)

    total = sum(
        interval["duration"] for intervals in grouped.values()
        for interval in intervals
    )
    rows = []
    for (machine, parent_state), intervals in sorted(grouped.items()):
        duration = sum(interval["duration"] for interval in intervals)
        rows.append({
            "machine": machine,
            "parent_state": parent_state,
            "count": len(intervals),
            "durations": [interval["duration"] for interval in intervals],
            "duration": duration,
            "ratio": duration / total if total else 0.0,
        })
    return rows


def analyze(events, window_start, window_end):
    log_start = events[0]["time"]
    log_end = events[-1]["time"]
    all_coverage = build_intervals(
        events, 1, "COVERAGE", log_start, log_end, infer_initial=True
    )
    coverage_intervals = clip_to_containers(
        all_coverage,
        [{
            "start": window_start, "end": window_end,
            "start_line": None, "end_line": None,
        }],
    )
    all_edge_coverage = build_intervals(
        events, 2, "EDGE_COVERAGE", log_start, log_end, infer_initial=True
    )
    edge_intervals = clip_to_containers(
        all_edge_coverage,
        coverage_intervals,
    )

    zpaths = build_intervals(events, 2, "ZPATH", window_start, window_end)
    if not zpaths:
        raise ValueError("指定时间段内没有有效的 HSM_2 ZPATH 状态")

    go_candidates = build_intervals(events, 3, "GO_POSITION", window_start, window_end)
    # GO_POSITION also exists under other HSM_2 states.  Only count the exact
    # portion where HSM_2 is ZPATH, so a child spanning a parent transition
    # cannot leak time from another parent state into the ZPATH statistics.
    go_intervals = clip_to_containers(go_candidates, zpaths)

    child_states = sorted(set(
        event["state"] for event in events
        if event["level"] == 4
        and window_start <= event["time"] <= window_end
    ))
    child_intervals = {}
    for state in child_states:
        candidates = build_intervals(events, 4, state, window_start, window_end)
        intervals = clip_to_containers(candidates, go_intervals)
        if intervals:
            child_intervals[state] = intervals

    rows = []
    for index, go in enumerate(go_intervals, 1):
        state_durations = {}
        for state, intervals in child_intervals.items():
            state_durations[state] = sum(
                intersection_seconds(go, interval) for interval in intervals
            )
        recorded_duration = sum(state_durations.values())
        rows.append({
            "index": index,
            "start": go["start"],
            "end": go["end"],
            "go_duration": go["duration"],
            "state_durations": state_durations,
            "other_duration": max(0.0, go["duration"] - recorded_duration),
            "start_line": go["start_line"],
            "end_line": go["end_line"],
        })
    return coverage_intervals, edge_intervals, zpaths, child_intervals, rows


def build_edge_traversals(rows, point_types, zpaths):
    """Classify every raw GO_POSITION interval without merging its duration.

    A point type is commonly emitted at the end of a retry chain.  We delay
    classification until that line is found, then assign its type to every raw
    GO_POSITION interval in the chain.  Each raw interval remains its own
    output record, so the category counts always add up to raw GO_POSITION.
    """
    traversals = []
    pending = []
    pending_zpath = None

    def zpath_index(row):
        return next(
            (index for index, interval in enumerate(zpaths)
             if intersection_seconds(row, interval) > 0),
            None,
        )

    def finish(point_type=None, point_line=None, completed=False):
        if not pending:
            return
        for row in pending:
            traversal_index = len(traversals) + 1
            item = {
                "index": traversal_index,
                "point_type": point_type,
                "edge_type": EDGE_TYPE_NAMES[point_type],
                "completed": completed,
                "start": row["start"],
                "end": row["end"],
                "go_duration": row["go_duration"],
                "wall_duration": (row["end"] - row["start"]).total_seconds(),
                "turn_duration": row["state_durations"].get("TURN", 0.0),
                "other_duration": row["other_duration"],
                "segment_count": 1,
                "start_line": row["start_line"],
                "end_line": row["end_line"],
                "point_line": point_line,
            }
            traversals.append(item)
            row["traversal_index"] = traversal_index
            row["point_type"] = point_type
            row["edge_type"] = item["edge_type"]
            row["traversal_completed"] = completed
        pending[:] = []

    for row in rows:
        current_zpath = zpath_index(row)
        if pending and current_zpath != pending_zpath:
            finish()
        if not pending:
            pending_zpath = current_zpath
        pending.append(row)

        start_line = row["start_line"]
        end_line = row["end_line"]
        hits = [] if start_line is None or end_line is None else [
            item for item in point_types
            if start_line < item["line"] <= end_line
        ]
        if hits:
            point = hits[-1]
            finish(
                point_type=point["point_type"],
                point_line=point["line"],
                completed=True,
            )

    finish()
    return traversals


def build_go_rows(events, window_start, window_end, zpaths,
                  infer_initial=False):
    """Build GO rows and level-4 child durations inside supplied ZPATHs."""
    go_candidates = build_intervals(
        events, 3, "GO_POSITION", window_start, window_end,
        infer_initial=infer_initial,
    )
    go_intervals = clip_to_containers(go_candidates, zpaths)
    child_states = sorted(set(
        event["state"] for event in events
        if event["level"] == 4
        and window_start <= event["time"] <= window_end
    ))
    child_intervals = {}
    for state in child_states:
        candidates = build_intervals(
            events, 4, state, window_start, window_end,
            infer_initial=infer_initial,
        )
        intervals = clip_to_containers(candidates, go_intervals)
        if intervals:
            child_intervals[state] = intervals

    rows = []
    for index, go in enumerate(go_intervals, 1):
        state_durations = {}
        for state, intervals in child_intervals.items():
            state_durations[state] = sum(
                intersection_seconds(go, interval) for interval in intervals
            )
        recorded_duration = sum(state_durations.values())
        rows.append({
            "index": index,
            "start": go["start"],
            "end": go["end"],
            "go_duration": go["duration"],
            "state_durations": state_durations,
            "other_duration": max(0.0, go["duration"] - recorded_duration),
            "start_line": go["start_line"],
            "end_line": go["end_line"],
        })
    return child_intervals, rows


def analyze_observed_task(events, point_types, raw_start, hsm_events,
                          cut_progress=None, focus_start=None,
                          focus_end=None, start_basis=None):
    """Summarize every observable COVERAGE session in this log.

    The first COVERAGE/ZPATH/GO state can be left-truncated when logging starts
    mid-task.  Such intervals are retained from the first raw timestamp and the
    summary explicitly reports that the resulting total is a lower bound.
    """
    observed_end = events[-1]["time"]
    coverage = build_intervals(
        events, 1, "COVERAGE", raw_start, observed_end, infer_initial=True
    )
    if not coverage:
        raise ValueError("日志内没有可观测的 COVERAGE 区间")
    all_periods = build_work_periods(coverage)
    next_period_start = None
    if focus_start and focus_end:
        focus_container = [{
            "start": focus_start, "end": focus_end,
            "start_line": None, "end_line": None,
        }]
        selected_indexes = [
            index for index, period in enumerate(all_periods)
            if period["start"] < focus_end and period["end"] > focus_start
        ]
        if selected_indexes:
            last_selected_index = selected_indexes[-1]
            if last_selected_index + 1 < len(all_periods):
                next_period_start = all_periods[last_selected_index + 1]["start"]
            coverage = clip_to_containers(coverage, focus_container)
    # The observable cycle is deliberately the wall-clock span bounded by the
    # outer COVERAGE intervals.  Do not substitute their accumulated runtime
    # for this total: recharge gaps are part of the cycle too.
    cycle_start = coverage[0]["start"]
    if focus_start and focus_start <= cycle_start:
        cycle_start = focus_start
    cycle_end = coverage[-1]["end"]
    cycle_total = (cycle_end - cycle_start).total_seconds()
    work_periods = build_work_periods(coverage)
    cut_progress = [
        item for item in (cut_progress or [])
        if cycle_start <= item["time"]
        and (next_period_start is None or item["time"] < next_period_start)
    ]

    edge = clip_to_containers(
        build_intervals(events, 2, "EDGE_COVERAGE", cycle_start, cycle_end,
                        infer_initial=True),
        coverage,
    )
    dock_events = [
        event for event in hsm_events if event["machine"] == "DOCK"
    ]
    dock_intervals = build_intervals(
        dock_events, 0, "DOCK", cycle_start, cycle_end, infer_initial=True
    )
    dock_goto_road = build_intervals(
        dock_events, 2, "GOTO_ROAD", cycle_start, cycle_end,
        infer_initial=True,
    )
    zpaths = clip_to_containers(
        build_intervals(events, 2, "ZPATH", cycle_start, cycle_end,
                        infer_initial=True),
        coverage,
    )
    child_intervals, rows = build_go_rows(
        events, cycle_start, cycle_end, zpaths, infer_initial=True
    )
    traversals = build_edge_traversals(rows, point_types, zpaths)

    recharge_gaps = []
    for previous, following in zip(coverage, coverage[1:]):
        if following["start"] > previous["end"]:
            recharge_gaps.append({
                "start": previous["end"],
                "end": following["start"],
                "duration": (following["start"] - previous["end"]).total_seconds(),
            })

    charging_intervals = []
    recharge_details = []
    for gap in recharge_gaps:
        def first_event(machine, action, state, start):
            return next((event["time"] for event in hsm_events
                         if event["machine"] == machine and event["level"] == 1
                         and event["action"] == action and event["state"] == state
                         and start <= event["time"] <= gap["end"]), None)

        dock_charge_inner = first_event("DOCK", "Inner", "DOCK_CHARGE", gap["start"])
        dock_charge_pop = first_event("DOCK", "Pop", "DOCK_CHARGE", gap["start"])
        cut_resume = next((event["time"] for event in hsm_events
                           if event["machine"] == "CUTTING" and event["level"] == 0
                           and event["action"] == "Init" and event["state"] == "CUT"
                           and dock_charge_pop is not None
                           and dock_charge_pop <= event["time"] <= gap["end"]), None)
        if dock_charge_pop and cut_resume:
            charging_intervals.append({
                "start": dock_charge_pop, "end": cut_resume,
                "duration": (cut_resume - dock_charge_pop).total_seconds(),
            })
            if dock_charge_inner:
                recharge_details.append({
                    "return_to_charger": (dock_charge_pop - gap["start"]).total_seconds(),
                    "dock_navigation": (dock_charge_inner - gap["start"]).total_seconds(),
                    "dock_alignment": (dock_charge_pop - dock_charge_inner).total_seconds(),
                    "charging": (cut_resume - dock_charge_pop).total_seconds(),
                    "resume_after_charge": (gap["end"] - cut_resume).total_seconds(),
                })
    # Recharge movement follows the dedicated HSM_0 DOCK state machine.  The
    # long electrical dwell starts only after DOCK is popped and is measured
    # independently in charging_intervals.
    recharge_operational_intervals = dock_intervals
    goto_road_by_parent = summarize_goto_road_by_parent(
        hsm_events, cycle_start, cycle_end
    )

    by_type = {
        edge_type: [item for item in traversals
                    if item["edge_type"] == edge_type]
        for edge_type in EDGE_TYPE_NAMES.values()
    }
    metrics = {
        "cycle_total": cycle_total,
        "coverage": sum(item["duration"] for item in coverage),
        "edge": sum(item["duration"] for item in edge),
        "recharge": sum(item["duration"] for item in dock_intervals),
        "charging": sum(item["duration"] for item in charging_intervals),
        "recharge_full_gap": sum(item["duration"] for item in recharge_gaps),
        "goto_breakpoint": sum(item["duration"] for item in dock_goto_road),
        # The task-level ZPATH metric counts only completed, explicitly typed
        # logical edges.  Parent-state planning/STOP time and INCOMPLETE rows
        # are excluded.
        "zpath": sum(
            item["go_duration"]
            for edge_type in ("LONG_EDGE", "SHORT_EDGE", "ASTAR")
            for item in by_type[edge_type]
        ),
        "short_edge": sum(item["go_duration"] for item in by_type["SHORT_EDGE"]),
        "long_edge": sum(item["go_duration"] for item in by_type["LONG_EDGE"]),
    }
    return {
        "start": cycle_start,
        "end": cycle_end,
        "coverage_intervals": coverage,
        "work_periods": work_periods,
        "edge_intervals": edge,
        "goto_intervals": dock_goto_road,
        "dock_intervals": dock_intervals,
        "zpaths": zpaths,
        "child_intervals": child_intervals,
        "rows": rows,
        "traversals": traversals,
        "recharge_gaps": recharge_gaps,
        "recharge_operational_intervals": recharge_operational_intervals,
        "charging_intervals": charging_intervals,
        "recharge_details": recharge_details,
        "goto_road_by_parent": goto_road_by_parent,
        "cut_progress": cut_progress,
        "metrics": metrics,
        "start_basis": start_basis,
        "long_count": len(by_type["LONG_EDGE"]),
        "short_count": len(by_type["SHORT_EDGE"]),
        "astar_count": len(by_type["ASTAR"]),
        "incomplete_count": len(by_type["INCOMPLETE"]),
        "raw_go_count": len(rows),
        "left_truncated": coverage[0]["start_line"] is None,
    }


def duration_text(seconds):
    milliseconds = int(round(seconds * 1000))
    hours, remainder = divmod(milliseconds, 3600000)
    minutes, remainder = divmod(remainder, 60000)
    secs, millis = divmod(remainder, 1000)
    return "{:02d}:{:02d}:{:02d}.{:03d}".format(hours, minutes, secs, millis)


def configure_datetime_axis(axis, mdates, start, end):
    locator = mdates.AutoDateLocator(minticks=4, maxticks=8)
    axis.set_xlim(start, end)
    axis.xaxis.set_major_locator(locator)
    axis.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))


def charging_removed_time_axis(start, end, charging_intervals):
    """Project datetimes onto an axis that collapses observed charging dwell."""
    intervals = []
    for interval in sorted(charging_intervals, key=lambda item: item["start"]):
        clipped_start = max(start, interval["start"])
        clipped_end = min(end, interval["end"])
        if clipped_start >= clipped_end:
            continue
        if intervals and clipped_start <= intervals[-1]["end"]:
            intervals[-1]["end"] = max(intervals[-1]["end"], clipped_end)
        else:
            intervals.append({"start": clipped_start, "end": clipped_end})

    visible = []
    cursor = start
    offset = 0.0
    for interval in intervals:
        if cursor < interval["start"]:
            duration = (interval["start"] - cursor).total_seconds()
            visible.append({
                "start": cursor,
                "end": interval["start"],
                "x_start": offset,
                "x_end": offset + duration,
            })
            offset += duration
        cursor = max(cursor, interval["end"])
    if cursor < end:
        duration = (end - cursor).total_seconds()
        visible.append({
            "start": cursor,
            "end": end,
            "x_start": offset,
            "x_end": offset + duration,
        })

    def project(value):
        if not visible or value <= visible[0]["start"]:
            return 0.0
        for segment in visible:
            if value <= segment["end"]:
                return segment["x_start"] + max(
                    0.0, (value - segment["start"]).total_seconds()
                )
        return visible[-1]["x_end"]

    tick_positions = []
    tick_labels = []
    for segment in visible:
        for value in (segment["start"], segment["end"]):
            position = project(value)
            if tick_positions and abs(position - tick_positions[-1]) < 0.001:
                # Preserve both endpoints at a collapsed charging gap while
                # keeping the charging dwell itself out of the axis length.
                tick_labels[-1] = "{} ->\n{}".format(
                    tick_labels[-1], value.strftime("%m-%d %H:%M"),
                )
                continue
            tick_positions.append(position)
            tick_labels.append(value.strftime("%m-%d %H:%M"))
    return project, visible, intervals, tick_positions, tick_labels


def configure_charging_removed_axis(axis, start, end, charging_intervals):
    """Configure an axis using absolute timestamps with charging gaps collapsed."""
    project, visible, removed, tick_positions, tick_labels = (
        charging_removed_time_axis(start, end, charging_intervals)
    )
    if visible:
        axis.set_xlim(0, visible[-1]["x_end"])
    axis.set_xticks(tick_positions)
    axis.set_xticklabels(tick_labels, fontsize=8)
    return project, visible, removed


def ensure_parent_dir(path):
    parent = Path(path).parent
    if str(parent) not in ("", "."):
        parent.mkdir(parents=True, exist_ok=True)


def output_path(path):
    """Place bare output filenames under outputs/ while respecting directories."""
    path = Path(path)
    if path.is_absolute() or str(path.parent) not in ("", "."):
        return path
    return DEFAULT_OUTPUT_DIR / path


def task_metric_segments(task):
    """Return per-COVERAGE and total-cycle metrics for chart display."""
    segments = []
    classified_edge_types = ("LONG_EDGE", "SHORT_EDGE", "ASTAR")

    def interval_total(intervals, container):
        return sum(intersection_seconds(interval, container) for interval in intervals)

    for index, coverage in enumerate(task["work_periods"], 1):
        segment_start = coverage["start"]
        if index == 1 and task["start"] < segment_start:
            segment_start = task["start"]
        zpath_total = sum(
            item["go_duration"] for item in task["traversals"]
            if item["completed"] and item["edge_type"] in classified_edge_types
            and intersection_seconds(item, coverage) > 0
        )
        segments.append({
            "label": "Period {}".format(index),
            "start": segment_start,
            "end": coverage["end"],
            "subtitle": "{} - {}".format(
                segment_start.strftime("%H:%M:%S"),
                coverage["end"].strftime("%H:%M:%S"),
            ),
            "metrics": {
                "cycle_total": (
                    coverage["end"] - segment_start
                ).total_seconds(),
                "coverage": coverage["coverage_duration"],
                "edge": interval_total(task["edge_intervals"], coverage),
                "recharge": interval_total(
                    task["recharge_operational_intervals"], coverage
                ),
                "goto_breakpoint": interval_total(
                    task["goto_intervals"], coverage
                ),
                "charging": interval_total(task["charging_intervals"], coverage),
                "zpath": zpath_total,
            },
        })

    segments.append({
        "label": "Total cycle",
        "start": task["start"],
        "end": task["end"],
        "subtitle": "{} - {}".format(
            task["start"].strftime("%H:%M:%S"),
            task["end"].strftime("%H:%M:%S"),
        ),
        "metrics": {
            key: task["metrics"][key] for key in TASK_METRIC_CHART_KEYS
        },
    })
    return segments


def build_work_periods(coverage_intervals):
    periods = []
    for interval in coverage_intervals:
        if not periods or (
                interval["start"] - periods[-1]["end"]
        ).total_seconds() > MERGE_COVERAGE_GAP_SECONDS:
            periods.append({
                "start": interval["start"],
                "end": interval["end"],
                "intervals": [interval],
            })
        else:
            periods[-1]["end"] = interval["end"]
            periods[-1]["intervals"].append(interval)

    for period in periods:
        period["wall_duration"] = (
            period["end"] - period["start"]
        ).total_seconds()
        period["coverage_duration"] = sum(
            interval["duration"] for interval in period["intervals"]
        )
        period["duration"] = period["wall_duration"]
    return periods


def observable_coverage_window(events, raw_start):
    coverage = build_intervals(
        events, 1, "COVERAGE", raw_start, events[-1]["time"],
        infer_initial=True,
    )
    if not coverage:
        raise ValueError("鏃ュ織鍐呮病鏈夊彲瑙傛祴鐨?COVERAGE 鍖洪棿")
    return coverage[0]["start"], coverage[-1]["end"]


def cut_zero_analysis_window(events, raw_start, cut_progress):
    zero = latest_true_zero_cut_progress(cut_progress)
    if not zero:
        start, end = observable_coverage_window(events, raw_start)
        return start, end, None

    coverage = build_intervals(
        events, 1, "COVERAGE", raw_start, events[-1]["time"],
        infer_initial=True,
    )
    later_coverage = [
        interval for interval in coverage if interval["end"] > zero["time"]
    ]
    if not later_coverage:
        raise ValueError("cut percent == 0 涔嬪悗娌℃湁鍙娴嬬殑 COVERAGE 鍖洪棿")
    return zero["time"], later_coverage[-1]["end"], zero


def describe(values):
    return {
        "count": len(values),
        "total": sum(values),
        "average": mean(values) if values else 0.0,
        "median": median(values) if values else 0.0,
        "minimum": min(values) if values else 0.0,
        "maximum": max(values) if values else 0.0,
    }


def print_summary(window_start, window_end, coverage_intervals, edge_intervals,
                  zpaths, child_intervals, rows, traversals):
    coverage_stats = describe([item["duration"] for item in coverage_intervals])
    edge_stats = describe([item["duration"] for item in edge_intervals])
    zpath_stats = describe([item["duration"] for item in zpaths])
    go_stats = describe([row["go_duration"] for row in rows])
    state_stats = {
        state: describe([item["duration"] for item in intervals])
        for state, intervals in child_intervals.items()
    }
    traversal_stats = {
        edge_type: describe([
            item["go_duration"] for item in traversals
            if item["edge_type"] == edge_type
        ])
        for edge_type in ("LONG_EDGE", "SHORT_EDGE", "ASTAR", "INCOMPLETE")
    }

    print("统计时间段: {} - {}".format(
        window_start.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
        window_end.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
    ))
    print("ZPATH 有效时长: {} ({:.3f}s)".format(
        duration_text(sum(item["duration"] for item in zpaths)),
        sum(item["duration"] for item in zpaths),
    ))
    print("{:<24} {:>6} {:>12} {:>10} {:>10} {:>10} {:>10}".format(
        "项目", "次数", "总时长(s)", "平均(s)", "中位(s)", "最小(s)", "最大(s)"
    ))
    summary_items = [
        ("COVERAGE", coverage_stats),
        ("COVERAGE 中 EDGE", edge_stats),
        ("ZPATH", zpath_stats),
        ("GO_POSITION", go_stats),
    ]
    summary_items.extend(
        ("GO 中 " + state, state_stats[state]) for state in sorted(state_stats)
    )
    summary_items.extend(
        (edge_type, traversal_stats[edge_type])
        for edge_type in ("LONG_EDGE", "SHORT_EDGE", "ASTAR", "INCOMPLETE")
        if traversal_stats[edge_type]["count"]
    )
    for label, stats in summary_items:
        print("{:<24} {:>6} {:>12.3f} {:>10.3f} {:>10.3f} {:>10.3f} {:>10.3f}".format(
            label, stats["count"], stats["total"], stats["average"],
            stats["median"], stats["minimum"], stats["maximum"]
        ))
    coverage_total = coverage_stats["total"]
    edge_ratio = edge_stats["total"] / coverage_total if coverage_total else 0.0
    print("EDGE_COVERAGE / COVERAGE 时长占比: {:.2%}".format(edge_ratio))
    zpath_total = sum(item["duration"] for item in zpaths)
    zpath_ratio = zpath_total / coverage_total if coverage_total else 0.0
    print("ZPATH / COVERAGE 时长占比: {:.2%}".format(zpath_ratio))
    go_ratio = go_stats["total"] / zpath_total if zpath_total else 0.0
    print("GO_POSITION / ZPATH 时长占比: {:.2%}".format(go_ratio))
    for state in sorted(state_stats):
        ratio = state_stats[state]["total"] / go_stats["total"] if go_stats["total"] else 0.0
        print("{} / GO_POSITION 时长占比: {:.2%}".format(state, ratio))
    completed = [item for item in traversals if item["completed"]]
    print("原始 GO 分段分类: 已分类 {} 段, 未完成 {} 段, 分类合计 {} = 原始 GO {}".format(
        len(completed), len(traversals) - len(completed),
        len(traversals), len(rows),
    ))


def export_csv(path, rows, child_states):
    ensure_parent_dir(path)
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.writer(output)
        header = ["序号", "GO开始", "GO结束", "GO时长(s)"]
        for state in child_states:
            header.extend([state + "时长(s)", state + "占比"])
        header.extend([
            "GO分类序号", "边类型", "point type", "是否完成",
            "开始行", "结束行",
        ])
        writer.writerow(header)
        for row in rows:
            values = [
                row["index"],
                row["start"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                row["end"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                "{:.3f}".format(row["go_duration"]),
            ]
            for state in child_states:
                duration = row["state_durations"].get(state, 0.0)
                ratio = duration / row["go_duration"] if row["go_duration"] else 0.0
                values.extend(["{:.3f}".format(duration), "{:.2%}".format(ratio)])
            values.extend([
                row.get("traversal_index"), row.get("edge_type"),
                row.get("point_type"), row.get("traversal_completed"),
                row["start_line"], row["end_line"],
            ])
            writer.writerow(values)


def export_excel(path, window_start, window_end, coverage_intervals,
                 edge_intervals, zpaths, child_intervals, rows, traversals,
                 task=None):
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Excel 导出需要 openpyxl，请执行: pip install openpyxl")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    details = workbook.create_sheet("GO明细")
    edge_details = workbook.create_sheet("边明细")
    state_details = workbook.create_sheet("状态明细")
    coverage_details = workbook.create_sheet("COVERAGE明细")
    task_metrics = workbook.create_sheet("任务指标") if task else None
    cut_progress_sheet = workbook.create_sheet("切割进度") if task else None
    task_period_header_row = None
    task_metric_header_row = None
    segments = task_metric_segments(task) if task else []
    if task:
        # Excel should use the same task window everywhere.
        window_start = task["start"]
        window_end = task["end"]
        coverage_intervals = task["coverage_intervals"]
        edge_intervals = task["edge_intervals"]
        zpaths = task["zpaths"]
        child_intervals = task["child_intervals"]
        rows = task["rows"]
        traversals = task["traversals"]
    child_states = sorted(child_intervals)
    coverage_stats = describe([item["duration"] for item in coverage_intervals])
    edge_stats = describe([item["duration"] for item in edge_intervals])
    zpath_stats = describe([item["duration"] for item in zpaths])
    go_stats = describe([row["go_duration"] for row in rows])
    zpath_total = zpath_stats["total"]

    summary.append(["ZPATH GO_POSITION 状态时长统计"])
    if task:
        summary.append(["ZPATH分析窗口开始", task["start"]])
        summary.append(["ZPATH分析窗口结束", task["end"]])
        if task.get("start_basis"):
            summary.append([
                "起点依据",
                "cut percent == 0（按 cut_area / total_area 高精度计算）",
                task["start_basis"]["time"],
                "cut_area", task["start_basis"]["cut_area"],
                "area_percent", task["start_basis"]["area_percent"],
                "logged_percent", task["start_basis"]["logged_percent"],
            ])
        for segment in segments[:-1]:
            total_seconds = segment["metrics"]["cycle_total"]
            summary.append([
                segment["label"], segment["start"], segment["end"],
                total_seconds, duration_text(total_seconds),
            ])
        summary.append(["总工作周期（含充电）", task["start"], task["end"],
                        task["metrics"]["cycle_total"],
                        duration_text(task["metrics"]["cycle_total"])])
        summary.append([
            "断点/回桩路径时间（DOCK.GOTO_ROAD）",
            task["metrics"]["goto_breakpoint"],
            duration_text(task["metrics"]["goto_breakpoint"]),
        ])
    else:
        summary.append(["ZPATH分析窗口开始", window_start])
        summary.append(["ZPATH分析窗口结束", window_end])
    summary.append(["ZPATH有效时长(s)", zpath_total])
    summary.append([])
    summary_header_row = summary.max_row + 1
    summary.append(["状态", "父状态", "次数", "总时长(s)", "平均(s)", "中位(s)",
                    "最小(s)", "最大(s)", "占父状态比例", "比例说明"])

    summary_rows = [
        ("COVERAGE", "COVERAGE", coverage_stats, coverage_stats["total"]),
    ]
    if task:
        dock_stats = describe([
            item["duration"] for item in task["dock_intervals"]
        ])
        summary_rows.append(
            ("DOCK（回桩）", "总工作周期", dock_stats,
             task["metrics"]["cycle_total"])
        )
        for item in task["goto_road_by_parent"]:
            if item["machine"] == "CUTTING" and item["parent_state"] == "COVERAGE":
                parent_total = coverage_stats["total"]
            elif item["machine"] == "DOCK":
                parent_total = dock_stats["total"]
            else:
                parent_total = task["metrics"]["cycle_total"]
            summary_rows.append((
                "GOTO_ROAD（{}）".format(item["machine"]),
                "{} / {}".format(item["machine"], item["parent_state"]),
                describe(item["durations"]), parent_total,
            ))
    summary_rows.extend([
        ("EDGE_COVERAGE", "COVERAGE", edge_stats, coverage_stats["total"]),
        ("ZPATH", "COVERAGE", zpath_stats, coverage_stats["total"]),
        ("GO_POSITION", "ZPATH", go_stats, zpath_total),
    ])
    summary_rows.extend(
        (state, "GO_POSITION",
         describe([item["duration"] for item in child_intervals[state]]),
         go_stats["total"])
        for state in child_states
    )
    summary_rows.extend(
        (edge_type, "GO_POSITION",
         describe([item["go_duration"] for item in traversals
                   if item["edge_type"] == edge_type]),
         go_stats["total"])
        for edge_type in ("LONG_EDGE", "SHORT_EDGE", "ASTAR", "INCOMPLETE")
        if any(item["edge_type"] == edge_type for item in traversals)
    )
    for state, parent_state, stats, parent_total in summary_rows:
        ratio = stats["total"] / parent_total if parent_total else 0.0
        summary.append([
            state, parent_state, stats["count"], stats["total"], stats["average"],
            stats["median"], stats["minimum"], stats["maximum"],
            ratio,
            "{} / {} 时长占比：{:.2%}".format(state, parent_state, ratio),
        ])
    summary_state_last_row = summary.max_row

    detail_header = ["序号", "GO开始", "GO结束", "GO时长(s)"]
    for state in child_states:
        detail_header.extend([state + "时长(s)", state + "占比"])
    detail_header.extend([
        "GO分类序号", "边类型", "point type", "是否完成", "开始行", "结束行"
    ])
    details.append(detail_header)
    for row in rows:
        values = [row["index"], row["start"], row["end"], row["go_duration"]]
        for state in child_states:
            duration = row["state_durations"].get(state, 0.0)
            ratio = duration / row["go_duration"] if row["go_duration"] else 0.0
            values.extend([duration, ratio])
        values.extend([
            row.get("traversal_index"), row.get("edge_type"),
            row.get("point_type"), row.get("traversal_completed"),
            row["start_line"], row["end_line"],
        ])
        details.append(values)

    edge_details.append([
        "GO分类序号", "边类型", "point type", "是否完成",
        "开始时间", "结束时间", "GO时长(s)", "墙钟时长(s)",
        "TURN时长(s)", "非TURN时长(s)", "GO分段数",
        "开始行", "结束行", "point type行",
    ])
    for item in traversals:
        edge_details.append([
            item["index"], item["edge_type"], item["point_type"],
            item["completed"], item["start"], item["end"],
            item["go_duration"], item["wall_duration"],
            item["turn_duration"], item["other_duration"],
            item["segment_count"], item["start_line"], item["end_line"],
            item["point_line"],
        ])

    state_details.append(["GO序号", "状态", "开始时间", "结束时间",
                          "时长(s)", "开始行", "结束行"])
    for state in child_states:
        for interval in child_intervals[state]:
            go_index = next(
                (row["index"] for row in rows
                 if intersection_seconds(row, interval) > 0),
                None,
            )
            state_details.append([
                go_index, state, interval["start"], interval["end"],
                interval["duration"], interval["start_line"], interval["end_line"],
            ])

    coverage_details.append([
        "类型", "序号", "开始时间", "结束时间", "时长(s)", "开始行", "结束行"
    ])
    for state, intervals in (
        ("COVERAGE", coverage_intervals),
        ("EDGE_COVERAGE", edge_intervals),
    ):
        for index, interval in enumerate(intervals, 1):
            coverage_details.append([
                state, index, interval["start"], interval["end"],
                interval["duration"], interval["start_line"], interval["end_line"],
            ])

    if task_metrics:
        bound_label = "可观测下限" if task["left_truncated"] else "完整可观测值"
        task_metrics.append(["日志内可观测覆盖任务指标"])
        start_label = "总工作周期开始"
        basis_text = "总时间（含充电） = 最后一个 COVERAGE end - 周期开始时间"
        if task.get("start_basis"):
            start_label += "（cut percent == 0）"
            basis_text = (
                "总时间（含充电） = 最后一个 COVERAGE end - cut percent == 0 时间点；"
                "cut percent 使用 cut_area / total_area 高精度计算"
            )
        else:
            start_label += "（首个 COVERAGE start）"
            basis_text = "总时间（含充电） = 最后一个 COVERAGE end - 首个 COVERAGE start"
        task_metrics.append([start_label, task["start"]])
        task_metrics.append(["总工作周期结束（最后一个 COVERAGE end）", task["end"]])
        task_metrics.append(["计算依据", basis_text])
        if task.get("start_basis"):
            task_metrics.append([
                "cut percent == 0 样本",
                task["start_basis"]["time"],
                "cut_area", task["start_basis"]["cut_area"],
                "total_area", task["start_basis"]["total_area"],
                "area_percent", task["start_basis"]["area_percent"],
                "logged_percent", task["start_basis"]["logged_percent"],
                "日志行", task["start_basis"]["line"],
            ])
        task_metrics.append(["统计性质", bound_label])
        task_metrics.append([])

        task_period_header_row = task_metrics.max_row + 1
        task_metrics.append([
            "时间段", "开始时间", "结束时间", "总时间（含充电）(s)",
            "总时间（含充电）", "统计性质",
        ])
        for segment in segments:
            total_seconds = segment["metrics"]["cycle_total"]
            task_metrics.append([
                segment["label"], segment["start"], segment["end"],
                total_seconds, duration_text(total_seconds), bound_label,
            ])
        task_metrics.append([])

        task_metric_header_row = task_metrics.max_row + 1
        second_headers = [segment["label"] + "(s)" for segment in segments]
        duration_headers = [segment["label"] for segment in segments]
        task_metrics.append(
            ["指标"] + second_headers + duration_headers + ["口径", "统计性质"]
        )
        for key, label, definition, _color in TASK_METRIC_DEFINITIONS:
            segment_seconds = [segment["metrics"][key] for segment in segments]
            task_metrics.append(
                [label] + segment_seconds
                + [duration_text(seconds) for seconds in segment_seconds]
                + [definition, bound_label]
            )

    if cut_progress_sheet:
        cut_progress_sheet.append(["切割面积和切割百分比"])
        cut_progress_sheet.append(["总工作周期开始", task["start"]])
        cut_progress_sheet.append(["总工作周期结束", task["end"]])
        cut_progress_sheet.append([])
        cut_progress_sheet.append([
            "时间", "已切割面积", "总面积", "切割百分比-面积计算(%)",
            "日志百分比(%)", "已切割时间(s)", "总计划时间(s)", "日志行",
        ])
        for item in task["cut_progress"]:
            cut_progress_sheet.append([
                item["time"], item["cut_area"], item["total_area"],
                item["area_percent"], item["logged_percent"],
                item["cut_time"], item["total_time"], item["line"],
            ])

    header_fill = PatternFill("solid", fgColor="176B87")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet, header_row in (
        (summary, summary_header_row), (details, 1), (edge_details, 1),
        (state_details, 1), (coverage_details, 1), (cut_progress_sheet, 5),
    ):
        if sheet is None or header_row is None:
            continue
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A{}".format(header_row + 1)
        sheet.auto_filter.ref = sheet.dimensions

    if task_metrics:
        for header_row in (task_period_header_row, task_metric_header_row):
            for cell in task_metrics[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        task_metrics.freeze_panes = "A{}".format(task_metric_header_row + 1)
        task_metric_max_column = 1 + len(segments) * 2 + 2
        task_metric_last_column = get_column_letter(task_metric_max_column)
        task_metrics.auto_filter.ref = "A{}:{}{}".format(
            task_metric_header_row, task_metric_last_column,
            task_metrics.max_row
        )
        task_metrics["A1"].font = Font(size=14, bold=True)
        task_metrics.merge_cells("A1:{}1".format(task_metric_last_column))
        task_metrics["A1"].alignment = Alignment(horizontal="center")
        for cell in (task_metrics["B2"], task_metrics["B3"]):
            cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
        for row_number in range(task_period_header_row + 1, task_metric_header_row - 1):
            task_metrics.cell(row_number, 2).number_format = "yyyy-mm-dd hh:mm:ss.000"
            task_metrics.cell(row_number, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
    if cut_progress_sheet:
        cut_progress_sheet["A1"].font = Font(size=14, bold=True)
        cut_progress_sheet.merge_cells("A1:H1")
        cut_progress_sheet["A1"].alignment = Alignment(horizontal="center")
        for cell in (cut_progress_sheet["B2"], cut_progress_sheet["B3"]):
            cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
        for row_number in range(6, cut_progress_sheet.max_row + 1):
            cut_progress_sheet.cell(row_number, 1).number_format = "yyyy-mm-dd hh:mm:ss.000"
            cut_progress_sheet.cell(row_number, 4).number_format = "0.00"
            cut_progress_sheet.cell(row_number, 5).number_format = "0.00"

    summary["A1"].font = Font(size=14, bold=True)
    summary.merge_cells("A1:J1")
    summary["A1"].alignment = Alignment(horizontal="center")
    if task:
        for row_number in range(3, 4 + len(segments[:-1])):
            summary.cell(row_number, 2).number_format = "yyyy-mm-dd hh:mm:ss.000"
            summary.cell(row_number, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
    else:
        for cell in (summary["B2"], summary["B3"]):
            cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
    for row_number in range(summary_header_row + 1, summary.max_row + 1):
        summary.cell(row_number, 9).number_format = "0.00%"
    for row_number in range(2, details.max_row + 1):
        details.cell(row_number, 2).number_format = "yyyy-mm-dd hh:mm:ss.000"
        details.cell(row_number, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
        for column in range(6, 4 + len(child_states) * 2, 2):
            details.cell(row_number, column).number_format = "0.00%"
    for row_number in range(2, state_details.max_row + 1):
        state_details.cell(row_number, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
        state_details.cell(row_number, 4).number_format = "yyyy-mm-dd hh:mm:ss.000"
    for row_number in range(2, edge_details.max_row + 1):
        edge_details.cell(row_number, 5).number_format = "yyyy-mm-dd hh:mm:ss.000"
        edge_details.cell(row_number, 6).number_format = "yyyy-mm-dd hh:mm:ss.000"
    for row_number in range(2, coverage_details.max_row + 1):
        coverage_details.cell(row_number, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
        coverage_details.cell(row_number, 4).number_format = "yyyy-mm-dd hh:mm:ss.000"

    widths = {
        summary: [20, 20, 12, 16, 14, 14, 14, 14, 16, 48],
        details: [8, 24, 24, 14] + [14, 12] * len(child_states) +
                 [12, 16, 12, 12, 10, 10],
        edge_details: [12, 16, 12, 12, 24, 24, 16, 16, 16, 18, 12, 10, 10, 12],
        state_details: [10, 20, 24, 24, 14, 10, 10],
        coverage_details: [20, 10, 24, 24, 14, 10, 10],
        cut_progress_sheet: [24, 14, 14, 22, 14, 16, 16, 10],
    }
    for sheet, column_widths in widths.items():
        if sheet is None:
            continue
        for index, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width
    if task_metrics:
        task_metric_widths = (
            [38] + [18] * len(segments) + [18] * len(segments) + [78, 16]
        )
        for index, width in enumerate(task_metric_widths, 1):
            task_metrics.column_dimensions[get_column_letter(index)].width = width

    if summary.max_row >= 7:
        chart = BarChart()
        chart.title = "各状态总时长"
        chart.y_axis.title = "时长(s)"
        data = Reference(summary, min_col=4, min_row=summary_header_row,
                         max_row=summary_state_last_row)
        categories = Reference(summary, min_col=1,
                               min_row=summary_header_row + 1,
                               max_row=summary_state_last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 13
        summary.add_chart(chart, "L2")

    if task_metrics:
        chart = BarChart()
        chart.title = "分时间段任务指标时长"
        chart.y_axis.title = "时长(s)"
        data = Reference(task_metrics, min_col=2, max_col=1 + len(segments),
                         min_row=task_metric_header_row,
                         max_row=task_metrics.max_row)
        categories = Reference(task_metrics, min_col=1,
                               min_row=task_metric_header_row + 1,
                               max_row=task_metrics.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.type = "bar"
        chart.style = 10
        chart.height = 8
        chart.width = 15
        task_metrics.add_chart(chart, "M2")

    if cut_progress_sheet and cut_progress_sheet.max_row >= 6:
        area_chart = LineChart()
        area_chart.title = "切割面积趋势"
        area_chart.y_axis.title = "面积"
        area_chart.x_axis.title = "时间"
        data = Reference(cut_progress_sheet, min_col=2, min_row=5,
                         max_row=cut_progress_sheet.max_row)
        categories = Reference(cut_progress_sheet, min_col=1, min_row=6,
                               max_row=cut_progress_sheet.max_row)
        area_chart.add_data(data, titles_from_data=True)
        area_chart.set_categories(categories)
        area_chart.height = 8
        area_chart.width = 15
        cut_progress_sheet.add_chart(area_chart, "I2")

        percent_chart = LineChart()
        percent_chart.title = "切割百分比趋势"
        percent_chart.y_axis.title = "百分比(%)"
        percent_chart.x_axis.title = "时间"
        data = Reference(cut_progress_sheet, min_col=4, max_col=5, min_row=5,
                         max_row=cut_progress_sheet.max_row)
        percent_chart.add_data(data, titles_from_data=True)
        percent_chart.set_categories(categories)
        percent_chart.height = 8
        percent_chart.width = 15
        cut_progress_sheet.add_chart(percent_chart, "I18")

    ensure_parent_dir(path)
    workbook.save(str(path))


def draw_chart(rows, window_start, window_end, coverage_intervals,
               edge_intervals, child_states, traversals,
               task=None, save_path=None, show=True):
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        raise RuntimeError("绘图需要 matplotlib，请执行: pip install matplotlib")

    indexes = [row["index"] for row in rows]
    go_values = [row["go_duration"] for row in rows]
    state_values = {
        state: [row["state_durations"].get(state, 0.0) for row in rows]
        for state in child_states
    }
    go_total = sum(go_values)
    state_totals = {
        state: sum(values) for state, values in state_values.items()
    }
    edge_values = {
        edge_type: [item["go_duration"] for item in traversals
                    if item["edge_type"] == edge_type]
        for edge_type in ("LONG_EDGE", "SHORT_EDGE", "ASTAR", "INCOMPLETE")
    }

    segment_count = len(task_metric_segments(task)) if task else 0
    crowded_segments = segment_count > 4
    fig_height = 8 if not crowded_segments else 9.5
    fig = plt.figure(figsize=(14, fig_height))
    grid = fig.add_gridspec(2, 2, height_ratios=[1.25, 1])
    timeline = fig.add_subplot(grid[0, :])
    totals = fig.add_subplot(grid[1, 0])
    distribution = fig.add_subplot(grid[1, 1])

    timeline.plot(indexes, go_values, color="#176B87", linewidth=1.2,
                  marker=".", markersize=3, label="GO_POSITION")
    colors = ["#D95F02", "#5E3C99", "#1B9E77", "#E6AB02", "#7570B3"]
    for index, state in enumerate(child_states):
        timeline.plot(
            indexes, state_values[state], color=colors[index % len(colors)],
            linewidth=1.0, marker=".", markersize=3,
            label="{} in GO_POSITION".format(state),
        )
    timeline.set_xlabel("GO_POSITION sequence")
    timeline.set_ylabel("Duration (seconds)")
    timeline.set_title("Duration by event")
    timeline.grid(axis="y", alpha=0.25)
    timeline.legend()

    if task:
        metric_labels = [
            chart_metric_label(key)
            for key in ZPATH_TASK_METRIC_CHART_KEYS
        ]
        segments = task_metric_segments(task)
        crowded_segments = len(segments) > 4
        group_gap = 1.0 if not crowded_segments else 1.35
        positions = [index * group_gap for index in range(len(metric_labels))]
        bar_height = min(0.22, 0.88 / max(1, len(segments)))
        segment_colors = ["#4C72B0", "#55A868", "#C44E52", "#8172B2"]

        for segment_index, segment in enumerate(segments):
            offset = (
                segment_index - (len(segments) - 1) / 2.0
            ) * bar_height
            values = [
                chart_metric_seconds(segment["metrics"], key) / 60.0
                for key in ZPATH_TASK_METRIC_CHART_KEYS
            ]
            y_positions = [position + offset for position in positions]
            totals.barh(
                y_positions, values, height=bar_height,
                color=segment_colors[segment_index % len(segment_colors)],
                label="{} ({})".format(segment["label"], segment["subtitle"]),
            )
            for key, y_position, value in zip(
                    ZPATH_TASK_METRIC_CHART_KEYS, y_positions, values):
                if value <= 0:
                    continue
                if crowded_segments:
                    if segment["label"] != "Total cycle" and key != "edge":
                        continue
                    label = "  {} {:.2f}m".format(
                        "Total" if segment["label"] == "Total cycle"
                        else segment["label"].replace("Period ", "P"),
                        value,
                    )
                else:
                    label = "  {:.2f}m {}".format(
                        value, duration_text(value * 60.0)
                    )
                totals.text(
                    value, y_position, label,
                    va="center", fontsize=6.0 if crowded_segments else 6.8,
                )

        totals.set_yticks(positions, metric_labels)
        totals.invert_yaxis()
        totals.set_xlabel("Duration (minutes)")
        totals.set_title(
            "Metrics by period and total cycle (charging excluded)"
            + (" (lower bounds)" if task["left_truncated"] else "")
        )
        totals.legend(loc="lower right", fontsize=7, frameon=True)
        totals.grid(axis="x", alpha=0.25)
        totals.margins(x=0.28)
    else:
        labels = ["COVERAGE", "EDGE_COVERAGE", "GO_POSITION"] + child_states + [
            "LONG_EDGE", "SHORT_EDGE"
        ]
        values = [
            sum(item["duration"] for item in coverage_intervals),
            sum(item["duration"] for item in edge_intervals),
            go_total,
        ] + [state_totals[state] for state in child_states] + [
            sum(edge_values["LONG_EDGE"]), sum(edge_values["SHORT_EDGE"])
        ]
        bar_colors = ["#4C4C4C", "#E6AB02", "#176B87"] + [
            colors[index % len(colors)] for index in range(len(child_states))
        ] + ["#C44E52", "#4C72B0"]
        totals.bar(labels, values, color=bar_colors)
        for index, value in enumerate(values):
            totals.text(index, value, "{:.3f}s".format(value), ha="center", va="bottom")
        totals.set_ylabel("Total duration (seconds)")
        totals.set_title("State duration totals")
        totals.grid(axis="y", alpha=0.25)
        totals.margins(y=0.12)
        for label in totals.get_xticklabels():
            label.set_rotation(20)
            label.set_horizontalalignment("right")

    classified_values = edge_values["LONG_EDGE"] + edge_values["SHORT_EDGE"]
    bins = min(30, max(8, int(len(classified_values) ** 0.5 * 2)))
    distribution.hist(edge_values["LONG_EDGE"], bins=bins,
                      color="#C44E52", alpha=0.68, label="LONG_EDGE")
    distribution.hist(edge_values["SHORT_EDGE"], bins=bins,
                      color="#4C72B0", alpha=0.68, label="SHORT_EDGE")
    distribution.set_xlabel("Duration (seconds)")
    distribution.set_ylabel("Count")
    distribution.set_title("Raw GO segment duration distribution")
    distribution.legend()
    distribution.grid(axis="y", alpha=0.25)

    if task:
        start_hint = "cut percent == 0" if task.get("start_basis") else "coverage"
        title = (
            "ZPATH timing analysis | {} task cycle {} - {} | analysis window {} - {}"
        ).format(
            start_hint,
            task["start"].strftime("%H:%M:%S.%f")[:-3],
            task["end"].strftime("%H:%M:%S.%f")[:-3],
            window_start.strftime("%H:%M:%S.%f")[:-3],
            window_end.strftime("%H:%M:%S.%f")[:-3],
        )
    else:
        title = "ZPATH timing analysis | {} - {}".format(
            window_start.strftime("%H:%M:%S.%f")[:-3],
            window_end.strftime("%H:%M:%S.%f")[:-3],
        )
    fig.suptitle(title, fontsize=14)
    fig.tight_layout()

    if save_path:
        ensure_parent_dir(save_path)
        fig.savefig(str(save_path), dpi=160, bbox_inches="tight")
        print("图表已保存: {}".format(save_path.resolve()))
    if show:
        plt.show()
    else:
        plt.close(fig)


def print_task_summary(task):
    bound_label = "（可观测下限）" if task["left_truncated"] else ""
    print("\n日志内可观测覆盖任务汇总:")
    print("  周期: {} - {}".format(
        task["start"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
        task["end"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
    ))
    if task.get("start_basis"):
        basis = task["start_basis"]
        print("  起点依据: cut percent == 0（按 cut_area / total_area 高精度计算）")
        print("    样本: {} cut_area={:.3f}, total_area={:.3f}, area_percent={:.6f}%, logged_percent={:.3f}%, line={}".format(
            basis["time"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            basis["cut_area"], basis["total_area"], basis["area_percent"],
            basis["logged_percent"], basis["line"],
        ))
        print("  计算依据: 总时间（含充电） = 最后一个 COVERAGE end - cut percent == 0 时间点")
    else:
        print("  计算依据: 总时间（含充电） = 最后一个 COVERAGE end - 首个 COVERAGE start")
    for key, label, _definition, _color in TASK_METRIC_DEFINITIONS:
        seconds = task["metrics"][key]
        print("  {:<42} {} ({:.3f}s){}".format(
            label, duration_text(seconds), seconds, bound_label
        ))
    print("  短边切割（含转弯） {} ({:.3f}s){}；长边切割（含转弯） {} ({:.3f}s){}".format(
        duration_text(task["metrics"]["short_edge"]), task["metrics"]["short_edge"], bound_label,
        duration_text(task["metrics"]["long_edge"]), task["metrics"]["long_edge"], bound_label,
    ))
    print("  原始 GO 分类 长/短/ASTAR/未完成: {}/{}/{}/{}；分类合计: {} = 原始 GO 段: {}".format(
        task["long_count"], task["short_count"], task["astar_count"],
        task["incomplete_count"], len(task["traversals"]), task["raw_go_count"],
    ))
    print("  GOTO_ROAD 按父状态（统计时间段内所有状态机）:")
    for item in task["goto_road_by_parent"]:
        print("    {} / {}: {} 次, {} ({:.3f}s, {:.2%})".format(
            item["machine"], item["parent_state"], item["count"],
            duration_text(item["duration"]), item["duration"], item["ratio"],
        ))
    for detail in task["recharge_details"]:
        print("  回充拆分: 返航入桩 {}（其中到充电点 {}、入桩校准 {}），"
              "充电驻留 {}（已排除），出桩恢复 {}".format(
                  duration_text(detail["return_to_charger"]),
                  duration_text(detail["dock_navigation"]),
                  duration_text(detail["dock_alignment"]),
                  duration_text(detail["charging"]),
                  duration_text(detail["resume_after_charge"]),
              ))
    if task["left_truncated"]:
        print("  注意: 日志从任务运行中途开始；上述周期总时间及所有累计统计均为可观测下限。")


def draw_task_summary(task, save_path):
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.lines import Line2D
        from matplotlib.patches import Patch
    except ImportError:
        raise RuntimeError("绘图需要 matplotlib，请执行: pip install matplotlib")

    metric_items = [
        item for item in TASK_METRIC_DEFINITIONS
        if item[0] in NON_CHARGING_TASK_METRIC_CHART_KEYS
    ]
    labels = [chart_metric_label(item[0]) for item in metric_items]
    values = [
        chart_metric_seconds(task["metrics"], item[0]) / 60.0
        for item in metric_items
    ]
    colors = [item[3] for item in metric_items]

    fig = plt.figure(figsize=(13, 14.4))
    grid = fig.add_gridspec(
        4, 1, height_ratios=[1, 1.45, 1.35, 1.45], hspace=0.68
    )
    timeline = fig.add_subplot(grid[0])
    bars = fig.add_subplot(grid[1])
    cut_axis = fig.add_subplot(grid[2])
    zpath_detail = fig.add_subplot(grid[3])

    state_colors = {
        "EDGE_COVERAGE": "#E6AB02",
        "GOTO_ROAD": "#55A868",
        "ZPATH": "#176B87",
        "OTHER": "#B8B8B8",
    }
    category_intervals = (
        ("EDGE_COVERAGE", task["edge_intervals"]),
        ("GOTO_ROAD", task["goto_intervals"]),
        ("ZPATH", task["zpaths"]),
    )
    task_start = task["start"]
    task_end = task["end"]
    project_task_time, _visible_task_time, _removed_charging = (
        configure_charging_removed_axis(
            timeline, task_start, task_end, task["charging_intervals"]
        )
    )

    def projected_duration(interval):
        return max(
            0.0,
            project_task_time(interval["end"]) - project_task_time(interval["start"]),
        )

    for period_index, period in enumerate(task["work_periods"], 1):
        # Paint each COVERAGE span as OTHER, then overlay its HSM_2 states.
        # GO_POSITION remains hidden because it is already contained by ZPATH.
        for coverage_interval in period["intervals"]:
            timeline.barh(
                0, projected_duration(coverage_interval),
                left=project_task_time(coverage_interval["start"]), height=0.42,
                color=state_colors["OTHER"],
            )
        for state, intervals in category_intervals:
            for interval in clip_to_containers(intervals, [period]):
                timeline.barh(
                    0, projected_duration(interval),
                    left=project_task_time(interval["start"]),
                    height=0.42, color=state_colors[state],
                )
        period_start = period["start"]
        if period_index == 1 and task_start < period_start:
            period_start = task_start
        midpoint = (
            project_task_time(period_start) + project_task_time(period["end"])
        ) / 2
        truncation = " (left-truncated)" if (
            period_index == 1 and task["left_truncated"]
        ) else ""
        timeline.text(
            midpoint, -0.30,
            "Period {}{}\n{}".format(
                period_index, truncation,
                duration_text((period["end"] - period_start).total_seconds()),
            ),
            ha="center", va="top", fontsize=8.5,
        )

    for interval in task["recharge_operational_intervals"]:
        timeline.barh(
            0, projected_duration(interval),
            left=project_task_time(interval["start"]),
            height=0.42, color="#8172B2",
        )
    for interval in _removed_charging:
        timeline.axvline(
            project_task_time(interval["start"]), color="#A0A0A0",
            linestyle="--", linewidth=0.8, alpha=0.75,
        )

    timeline.set_ylim(-0.62, 0.48)
    timeline.set_yticks([])
    timeline.set_xlabel("Log time (charging collapsed; discontinuous)")
    if task.get("start_basis"):
        timeline_title = (
            "Observable cycle: cut percent == 0 to last COVERAGE end "
            "(charging collapsed)"
        )
    else:
        timeline_title = (
            "Observable cycle: first COVERAGE start to last COVERAGE end "
            "(charging collapsed)"
        )
    timeline.set_title(timeline_title)
    timeline.legend(
        handles=[
            Patch(color=state_colors["EDGE_COVERAGE"], label="Edge coverage"),
            Patch(color=state_colors["GOTO_ROAD"],
                  label="GOTO_ROAD / Go to breakpoint"),
            Patch(color=state_colors["ZPATH"],
                  label="ZPATH (includes GO_POSITION)"),
            Patch(color=state_colors["OTHER"], label="Other active time"),
            Patch(color="#8172B2", label="Recharge overhead"),
            Line2D(
                [0], [0], color="#A0A0A0", linestyle="--",
                label="Charging (collapsed: start -> end)",
            ),
        ],
        loc="lower center", bbox_to_anchor=(0.5, 1.25),
        ncol=3, frameon=False, fontsize=8.5,
    )
    timeline.grid(axis="x", alpha=0.22)

    positions = list(range(len(labels)))
    bars.barh(positions, values, color=colors)
    bars.set_yticks(positions, labels)
    bars.invert_yaxis()
    bars.set_xlabel("Duration (minutes)")
    bars.set_title("Task timing metrics (charging excluded)")
    bars.grid(axis="x", alpha=0.22)
    bars.margins(x=0.16)
    metric_keys = [item[0] for item in metric_items]
    for position, key, value in zip(positions, metric_keys, values):
        seconds = value * 60.0
        bars.text(
            value, position,
            "  {:.2f} min  ({})".format(value, duration_text(seconds)),
            va="center", fontsize=10,
        )

    progress = task.get("cut_progress", [])
    if progress:
        cut_times = [item["time"] for item in progress]
        cut_areas = [item["cut_area"] for item in progress]
        total_areas = [item["total_area"] for item in progress]
        area_percents = [item["area_percent"] for item in progress]
        logged_percents = [item["logged_percent"] for item in progress]
        progress_end = max(task_end, cut_times[-1])
        cut_percent_axis = cut_axis.twinx()
        project_cut_time, _visible_cut_time, _removed_cut_charging = (
            configure_charging_removed_axis(
                cut_axis, task_start, progress_end, task["charging_intervals"]
            )
        )
        for period_index, period in enumerate(task["work_periods"], 1):
            cut_axis.axvspan(
                project_cut_time(period["start"]), project_cut_time(period["end"]),
                color="#F1F3F5" if period_index % 2 else "#E8F2F0",
                alpha=0.8, zorder=0,
            )
        for interval in _removed_cut_charging:
            cut_axis.axvline(
                project_cut_time(interval["start"]), color="#A0A0A0",
                linestyle="--", linewidth=0.8, alpha=0.7, zorder=1,
            )
        cut_axis.plot(
            [project_cut_time(value) for value in cut_times],
            cut_areas, color="#176B87", linewidth=1.5,
            marker=".", markersize=2.5, label="Cut area",
        )
        cut_axis.plot(
            [project_cut_time(value) for value in cut_times],
            total_areas, color="#8A8A8A", linewidth=0.9,
            linestyle="--", label="Total area",
        )
        cut_percent_axis.plot(
            [project_cut_time(value) for value in cut_times],
            area_percents, color="#C44E52", linewidth=1.4,
            marker=".", markersize=2.5, label="Cut percent (area calc)",
        )
        cut_percent_axis.plot(
            [project_cut_time(value) for value in cut_times],
            logged_percents, color="#C44E52", linewidth=0.8,
            linestyle=":", alpha=0.55, label="Logged percent",
        )
        cut_axis.set_xlabel("Log time (charging collapsed; discontinuous)")
        cut_axis.set_ylabel("Area")
        cut_percent_axis.set_ylabel("Percent (%)")
        cut_percent_axis.set_ylim(0, max(100, max(area_percents) * 1.08))
        final = progress[-1]
        cut_axis.set_title(
            "Cut progress (charging collapsed) | final {:.2f}/{:.2f}, {:.2f}%".format(
                final["cut_area"], final["total_area"],
                final["area_percent"],
            )
        )
        cut_axis.grid(axis="both", alpha=0.22)
        lines, line_labels = cut_axis.get_legend_handles_labels()
        right_lines, right_labels = cut_percent_axis.get_legend_handles_labels()
        cut_axis.legend(
            lines + right_lines, line_labels + right_labels,
            loc="upper left", frameon=False, ncol=4, fontsize=8.5,
        )
    else:
        cut_axis.text(
            0.5, 0.5, "No cut progress records in task cycle",
            transform=cut_axis.transAxes, ha="center", va="center",
            fontsize=11,
        )
        cut_axis.set_axis_off()

    long_items = [
        item for item in task["traversals"]
        if item["edge_type"] == "LONG_EDGE"
    ]
    short_items = [
        item for item in task["traversals"]
        if item["edge_type"] == "SHORT_EDGE"
    ]
    astar_items = [
        item for item in task["traversals"]
        if item["edge_type"] == "ASTAR"
    ]
    project_zpath_time, visible_zpath_time, removed_charging, tick_positions, tick_labels = (
        charging_removed_time_axis(
            task_start, task_end, task["charging_intervals"]
        )
    )
    for items, color, label, marker in (
        (long_items, "#C44E52", "Long edge", "o"),
        (short_items, "#4C72B0", "Short edge", "s"),
        (astar_items, "#09FF00", "ASTAR", "^"),
    ):
        start_times = [project_zpath_time(item["start"]) for item in items]
        durations = [item["go_duration"] for item in items]
        zpath_detail.vlines(
            start_times, 0, durations, color=color, alpha=0.20,
            linewidth=0.7,
        )
        zpath_detail.scatter(
            start_times, durations, color=color, s=15, marker=marker,
            label="{}: {} traversals, {}".format(
                label, len(items),
                duration_text(sum(durations)),
            ),
        )
    for interval in removed_charging:
        boundary = project_zpath_time(interval["start"])
        zpath_detail.axvline(
            boundary, color="#A0A0A0", linestyle="--", linewidth=0.8,
            alpha=0.7,
        )
    if visible_zpath_time:
        zpath_detail.set_xlim(0, visible_zpath_time[-1]["x_end"])
    zpath_detail.set_xticks(tick_positions)
    zpath_detail.set_xticklabels(tick_labels, fontsize=8)
    zpath_detail.set_xlabel("Log time (charging collapsed; discontinuous)")
    zpath_detail.set_ylabel("Duration (seconds)")
    zpath_detail.set_title(
        "ZPATH detail by absolute log time (charging collapsed)", pad=30,
    )
    zpath_detail.grid(axis="y", alpha=0.22)
    zpath_detail.legend(
        loc="upper center", bbox_to_anchor=(0.5, 1.02),
        frameon=False, ncol=3,
    )

    note = (
        "Raw GO classification: long {} | short {} | ASTAR {} | incomplete {}"
        " | total {} = raw GO segments {}"
    ).format(
        task["long_count"], task["short_count"], task["astar_count"],
        task["incomplete_count"], len(task["traversals"]), task["raw_go_count"],
    )
    if task["left_truncated"]:
        note += "\nLog starts mid-task; cycle total and all cumulative statistics are observable lower bounds."
    fig.text(0.5, 0.015, note, ha="center", fontsize=10)
    ensure_parent_dir(save_path)
    fig.savefig(str(save_path), dpi=170, bbox_inches="tight")
    plt.close(fig)
    print("任务总览图已保存: {}".format(save_path.resolve()))


def draw_cut_progress(task, save_path):
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        raise RuntimeError("绘图需要 matplotlib，请执行: pip install matplotlib")

    progress = task.get("cut_progress", [])
    if not progress:
        print("切割进度图跳过: 当前任务周期内没有 cut info 记录")
        return

    times = [item["time"] for item in progress]
    cut_areas = [item["cut_area"] for item in progress]
    total_areas = [item["total_area"] for item in progress]
    percents = [item["area_percent"] for item in progress]
    logged_percents = [item["logged_percent"] for item in progress]

    fig, area_axis = plt.subplots(figsize=(13, 6.2))
    percent_axis = area_axis.twinx()
    progress_end = max(task["end"], times[-1])
    project_progress_time, _visible_progress_time, _removed_progress_charging = (
        configure_charging_removed_axis(
            area_axis, task["start"], progress_end,
            task["charging_intervals"],
        )
    )

    for index, period in enumerate(task["work_periods"], 1):
        area_axis.axvspan(
            project_progress_time(period["start"]),
            project_progress_time(period["end"]),
            color="#F1F3F5" if index % 2 else "#E8F2F0",
            alpha=0.8, zorder=0,
        )
        midpoint = (
            project_progress_time(period["start"])
            + project_progress_time(period["end"])
        ) / 2
        area_axis.text(
            midpoint, 0.02, "Period {}".format(index),
            transform=area_axis.get_xaxis_transform(),
            ha="center", va="bottom", fontsize=9, color="#4A4A4A",
        )
    for interval in _removed_progress_charging:
        area_axis.axvline(
            project_progress_time(interval["start"]), color="#A0A0A0",
            linestyle="--", linewidth=0.8, alpha=0.7, zorder=1,
        )

    area_axis.plot(
        [project_progress_time(value) for value in times],
        cut_areas, color="#176B87", linewidth=1.8,
        marker=".", markersize=3, label="Cut area",
    )
    area_axis.plot(
        [project_progress_time(value) for value in times],
        total_areas, color="#8A8A8A", linewidth=1.0,
        linestyle="--", label="Total area",
    )
    percent_axis.plot(
        [project_progress_time(value) for value in times],
        percents, color="#C44E52", linewidth=1.6,
        marker=".", markersize=3, label="Cut percent (area calc)",
    )
    percent_axis.plot(
        [project_progress_time(value) for value in times],
        logged_percents, color="#C44E52", linewidth=0.9,
        linestyle=":", alpha=0.55, label="Logged percent",
    )

    area_axis.set_xlabel(
        "Absolute log time (charging intervals collapsed; discontinuous)"
    )
    area_axis.set_ylabel("Area")
    percent_axis.set_ylabel("Percent (%)")
    percent_axis.set_ylim(0, max(100, max(percents) * 1.08))
    area_axis.grid(axis="both", alpha=0.22)

    final = progress[-1]
    area_axis.set_title(
        "Cut area and percent (charging collapsed) | final {:.2f}/{:.2f}, {:.1f}%".format(
            final["cut_area"], final["total_area"], final["area_percent"],
        )
    )
    lines, labels = area_axis.get_legend_handles_labels()
    right_lines, right_labels = percent_axis.get_legend_handles_labels()
    area_axis.legend(
        lines + right_lines, labels + right_labels,
        loc="upper left", frameon=False, ncol=3,
    )

    fig.tight_layout()
    ensure_parent_dir(save_path)
    fig.savefig(str(save_path), dpi=170, bbox_inches="tight")
    plt.close(fig)
    print("切割进度图已保存: {}".format(save_path.resolve()))


def main():
    args = parse_args()
    log_path = Path(args.log)
    records = load_log_records(log_path)
    events = parse_log(records)
    point_types = parse_point_types(records)
    cut_progress = parse_cut_progress(records)
    raw_start, hsm_events = parse_raw_context(records)
    print("日志输入: {} ({} files, {} timestamped lines)".format(
        log_path,
        len(log_files(log_path)),
        len(records),
    ))
    cut_zero_start = None
    if args.start == DEFAULT_START and args.end == DEFAULT_END:
        window_start, window_end, cut_zero_start = cut_zero_analysis_window(
            events, raw_start, cut_progress
        )
        if cut_zero_start:
            print(
                "分析起点: cut percent == 0 @ {} "
                "(cut_area={:.3f}, total_area={:.3f}, area_percent={:.6f}%, "
                "logged_percent={:.3f}%, line={})".format(
                    cut_zero_start["time"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                    cut_zero_start["cut_area"],
                    cut_zero_start["total_area"],
                    cut_zero_start["area_percent"],
                    cut_zero_start["logged_percent"],
                    cut_zero_start["line"],
                )
            )
        result = analyze(events, window_start, window_end)
    else:
        selected = None
        last_error = None
        for window_start, window_end in candidate_windows(args.start, args.end, events):
            try:
                result = analyze(events, window_start, window_end)
            except ValueError as error:
                last_error = error
                continue
            if result[-1]:
                selected = (window_start, window_end, result)
        if selected is None:
            if last_error:
                raise last_error
            raise ValueError("指定时间段的 ZPATH 中没有找到 GO_POSITION")
        window_start, window_end, result = selected
    coverage_intervals, edge_intervals, zpaths, child_intervals, rows = result

    traversals = build_edge_traversals(rows, point_types, zpaths)
    # Build once so Excel, console, and task chart share the exact same
    # observable-cycle boundaries and metric definitions.
    task = analyze_observed_task(
        events, point_types, raw_start, hsm_events,
        cut_progress=cut_progress,
        focus_start=window_start, focus_end=window_end,
        start_basis=cut_zero_start,
    )

    child_states = sorted(child_intervals)
    print_summary(
        window_start, window_end, coverage_intervals, edge_intervals,
        zpaths, child_intervals, rows, traversals,
    )
    if args.csv:
        csv_path = output_path(args.csv)
        export_csv(csv_path, rows, child_states)
        print("明细已导出: {}".format(csv_path.resolve()))
    if args.excel:
        excel_path = output_path(args.excel)
        export_excel(
            excel_path, window_start, window_end,
            coverage_intervals, edge_intervals, zpaths,
            child_intervals, rows, traversals, task,
        )
        print("Excel 已导出: {}".format(excel_path.resolve()))
    chart_path = output_path(args.save) if args.save else None
    draw_chart(
        rows, window_start, window_end, coverage_intervals,
        edge_intervals, child_states, traversals,
        task=task,
        save_path=chart_path,
        show=not args.no_show,
    )
    if chart_path:
        chart_alias = DEFAULT_CHART_SAVE
        if chart_path.resolve() != chart_alias.resolve():
            ensure_parent_dir(chart_alias)
            shutil.copyfile(str(chart_path), str(chart_alias))
            print("图表副本已保存: {}".format(chart_alias.resolve()))

    if args.task_summary:
        print_task_summary(task)
    if args.task_save:
        draw_task_summary(task, output_path(args.task_save))
    if args.cut_save:
        draw_cut_progress(task, output_path(args.cut_save))


if __name__ == "__main__":
    try:
        main()
    except (OSError, ValueError, RuntimeError) as error:
        print("错误: {}".format(error), file=sys.stderr)
        sys.exit(1)
